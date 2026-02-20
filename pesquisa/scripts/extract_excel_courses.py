#!/usr/bin/env python3
"""
Extract course names from Excel planning file
Usage: python3 extract_excel_courses.py <filepath>
"""

import sys
import json
import openpyxl

def extract_courses(filepath):
    """Extract course names from Excel file"""
    try:
        # Load workbook
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        courses = []
        skip_keywords = ['CUSTO', 'HORAS AULA', 'TOTAL', 'SUBTOTAL', 'SOMA']
        
        # Start from row 4, column D (index 4)
        for row in range(4, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=4).value
            
            if cell_value and isinstance(cell_value, str):
                course_name = cell_value.strip()
                
                # Skip if too short
                if len(course_name) < 5:
                    continue
                
                # Skip if contains skip keywords
                if any(keyword in course_name.upper() for keyword in skip_keywords):
                    continue
                
                courses.append(course_name)
        
        return courses
        
    except Exception as e:
        print(json.dumps({'error': str(e)}), file=sys.stderr)
        return []

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print(json.dumps({'error': 'Usage: python3 extract_excel_courses.py <filepath>'}))
        sys.exit(1)
    
    filepath = sys.argv[1]
    courses = extract_courses(filepath)
    
    # Output as JSON
    print(json.dumps(courses, ensure_ascii=False))